#!/usr/bin/env python3
"""Escribir ratings y fórmula de promedio en el Excel del Torneo ACI.

Uso:
  1) Corre scrape.js -> genera profiles.json
  2) Ajusta EXCEL abajo si cambia el nombre del archivo.
  3) python3 write_excel.py            # lee profiles.json y escribe D:G + H
     python3 write_excel.py --check    # solo imprime lo que haría (no guarda)

Notas:
- openpyxl NO evalúa fórmulas; H se calcula al abrir en Excel/Google Sheets.
  Este script también imprime el promedio calculado en Python para verificar.
- Filas sin 1v1 (elo_1v1_act = null) usan la fórmula de promedio solo-Team renormalizada.
- Requisitos: pip install openpyxl   (idealmente en un venv)
"""
import json
import os
import sys

import openpyxl

EXCEL = "Inscripcion Torneo ACI (Respuestas).xlsx"
PROFILES = "profiles.json"

# Pesos (deben coincidir con la hoja Notas!B3:B6). Solo para el cálculo de verificación.
W1, W2, W3, W4 = 0.35, 0.30, 0.20, 0.15  # 1v1 act, team act, 1v1 max, team max


def h_formula(row: int, has_1v1: bool) -> str:
    if has_1v1:
        return (f"=ROUND(D{row}*Notas!$B$3+E{row}*Notas!$B$4"
                f"+F{row}*Notas!$B$5+G{row}*Notas!$B$6,0)")
    # sin 1v1: promedio solo con Team, renormalizado
    return f"=ROUND((E{row}*Notas!$B$4+G{row}*Notas!$B$6)/(Notas!$B$4+Notas!$B$6),0)"


def h_value(d, e, f, g):
    if d is None:
        return round((e * W2 + g * W4) / (W2 + W4))
    return round(d * W1 + e * W2 + f * W3 + g * W4)


def main():
    check = "--check" in sys.argv
    if not os.path.exists(PROFILES):
        sys.exit(f"No existe {PROFILES}. Corre primero scrape.js.")
    data = json.load(open(PROFILES))

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    print(f"{'Fila':>4}  {'D':>5} {'E':>5} {'F':>5} {'G':>5}  {'H(calc)':>7}")
    for row_str, p in data.items():
        if "error" in p:
            print(f"{row_str:>4}  ERROR: {p['error']}")
            continue
        r = int(row_str)
        d = p.get("elo_1v1_act")
        e = p.get("elo_team_act")
        f = p.get("elo_1v1_max")
        g = p.get("elo_team_max")
        h = h_value(d, e, f, g)
        print(f"{r:>4}  {str(d or '-'):>5} {str(e or '-'):>5} "
              f"{str(f or '-'):>5} {str(g or '-'):>5}  {h:>7}")
        if not check:
            ws.cell(r, 4).value = d
            ws.cell(r, 5).value = e
            ws.cell(r, 6).value = f
            ws.cell(r, 7).value = g
            ws.cell(r, 8).value = h_formula(r, d is not None)
    if check:
        print("\n--check: no se guardó nada.")
    else:
        wb.save(EXCEL)
        print(f"\nGuardado en {EXCEL}. H se recalcula al abrir en Excel/Sheets.")


if __name__ == "__main__":
    main()
